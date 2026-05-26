#!/usr/bin/env python3
"""
TIA Portal HMI Full Data Extractor
====================================
Combines three data sources into a comprehensive per-element extraction:

  1. RDF screen files    → element names, types, JS events, navigation, PLC tag refs
  2. HMITags.xlsx        → HMI tag name ↔ PLC tag mapping, data types, addresses
  3. OCR data (optional) → element positions, text labels, colors from screenshots

Usage:
    python extract_hmi_full.py "<TIA_PROJECT_PATH>" [options]

Options:
    --instance ID     HMI instance ID (auto-detects latest if omitted)
    --output PATH     Output JSON file (default: Doc_OUTPUT/hmi_offline_data.json)
    --tags FILE       HMITags.xlsx path (default: DATA_HMI/HMITags.xlsx)
    --ocr FILE        OCR all_elements.json path (default: Doc_OUTPUT/all_elements.json)
    --list-hmis       List all HMI instances and exit
    --verbose         Print detailed progress
"""

import os
import re
import sys
import json
import argparse
import struct
from collections import defaultdict
from datetime import datetime

# ---------------------------------------------------------------------------
# HMITags.xlsx reader (requires openpyxl)
# ---------------------------------------------------------------------------
def load_hmi_tags(xlsx_path):
    """Load HMITags.xlsx and return dict of HMI_tag_name -> tag_details."""
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed. Run: pip install openpyxl")
        return {}

    if not os.path.exists(xlsx_path):
        print(f"WARNING: HMITags.xlsx not found at {xlsx_path}")
        return {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    tags = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else '' for h in rows[0]]

        for row in rows[1:]:
            values = list(row)
            row_dict = {}
            for i, h in enumerate(headers):
                if i < len(values):
                    row_dict[h] = values[i] if values[i] is not None else ''

            name = row_dict.get('Name', '')
            if not name:
                continue

            tags[name] = {
                'hmi_tag': name,
                'path': str(row_dict.get('Path', '')),
                'connection': str(row_dict.get('Connection', '')),
                'plc_tag': str(row_dict.get('PLC tag', '')),
                'data_type': str(row_dict.get('DataType', '')),
                'hmi_data_type': str(row_dict.get('HMI DataType', '')),
                'length': row_dict.get('Length', ''),
                'access_method': str(row_dict.get('Access Method', '')),
                'address': str(row_dict.get('Address', '')),
                'start_value': str(row_dict.get('Start value', '')),
                'comment': str(row_dict.get('Comment [en-US]', '')),
                'acquisition_mode': str(row_dict.get('Acquisition mode', '')),
                'acquisition_cycle': str(row_dict.get('Acquisition cycle', '')),
                'limit_upper': str(row_dict.get('Limit Upper 2', '')),
                'limit_lower': str(row_dict.get('Limit Lower 2', '')),
            }

    wb.close()
    print(f"Loaded {len(tags)} HMI tags from {os.path.basename(xlsx_path)}")
    return tags


# ---------------------------------------------------------------------------
# OCR data loader
# ---------------------------------------------------------------------------
def load_ocr_data(json_path):
    """Load OCR all_elements.json and return dict of screen_filename -> elements."""
    if not os.path.exists(json_path):
        print(f"WARNING: OCR data not found at {json_path}")
        return {}

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    screens = {}
    for filename, screen_data in data.get('screens', {}).items():
        # Normalize filename to screen name
        base = os.path.splitext(filename)[0]
        elements = []
        for elem in screen_data.get('elements', []):
            elements.append({
                'text': elem.get('text', ''),
                'type_ocr': elem.get('type', ''),
                'bbox': elem.get('bbox', {}),
                'center': elem.get('center', {}),
                'confidence': elem.get('confidence', 0),
                'color_region': elem.get('color_region'),
            })
        screens[base] = elements

    print(f"Loaded OCR data for {len(screens)} screens from {os.path.basename(json_path)}")
    return screens


# ---------------------------------------------------------------------------
# RDF binary parser - extract element details
# ---------------------------------------------------------------------------
def extract_js_functions(data):
    """Extract JavaScript function definitions from RDF binary data."""
    js_pattern = rb'export\s+async\s+function\s+(\w+)\s*\([^)]*\)\s*\{'
    functions = []

    for match in re.finditer(js_pattern, data):
        func_name = match.group(1).decode('utf-8', errors='ignore')
        brace_count = 0
        func_start = data.index(b'{', match.start())
        pos = func_start
        while pos < len(data) and pos - func_start < 10000:
            if data[pos:pos+1] == b'{':
                brace_count += 1
            elif data[pos:pos+1] == b'}':
                brace_count -= 1
                if brace_count == 0:
                    func_body = data[match.start():pos+1].decode('utf-8', errors='ignore')
                    functions.append({'name': func_name, 'body': func_body})
                    break
            pos += 1
    return functions


def extract_plc_tags(js_code):
    """Extract PLC tag references from JavaScript code."""
    tags = set()
    patterns = [
        r'GetTagValue\("([^"]+)"\)',
        r'SetTagValue\("([^"]+)"',
        r'SetBitInTag\("([^"]+)"',
        r'ResetBitInTag\("([^"]+)"',
        r'TagValue\("([^"]+)"\)',
        r'"(DB_[A-Z_0-9][A-Z_0-9 .\-]+)"',
        r'"(HOPPER_[A-Z_0-9]+)"',
        r'"(DataToerenw_[A-Z_0-9]+)"',
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, js_code):
            tags.add(match.group(1).strip())
    return sorted(tags)


def extract_screen_navigation(js_code):
    """Extract screen navigation calls."""
    navigations = []
    for match in re.finditer(r'ChangeScreen\("([^"]+)"', js_code):
        navigations.append(match.group(1))
    return navigations


def extract_element_details(data):
    """Extract UI elements with their binary property data from RDF."""
    elements = []
    # Pattern: element name followed by binary properties
    elem_pattern = rb'((?:Button|Text box|Screen window|IO field|Bar|Symbolic IO field|Status display|Switch|Slider|Clock|Date|Text list|Image|Rectangle|Line|Circle|Faceplate)_(?:\w+))'

    # Event suffixes to exclude - these are JS function names, not element names
    event_suffixes = ('_OnUp', '_OnDown', '_OnClick', '_OnChange', '_OnLoaded',
                      '_OnMouseEnter', '_OnMouseLeave', '_OnFocus', '_OnBlur',
                      '_OnKeyPress', '_OnKeyDown', '_OnKeyUp', '_OnShow', '_OnHide')

    for match in re.finditer(elem_pattern, data):
        elem_name = match.group(1).decode('utf-8', errors='ignore')
        elem_start = match.start()

        # Skip JS function names (e.g. Button_9_OnUp is a function, not an element)
        if any(elem_name.endswith(s) for s in event_suffixes):
            continue

        if elem_name in [e['name'] for e in elements]:
            continue

        # Determine element type from name
        elem_type = 'unknown'
        type_map = {
            'Button': 'button',
            'Text box': 'text_display',
            'Screen window': 'screen_window',
            'IO field': 'io_field',
            'Bar': 'bar_graph',
            'Symbolic IO field': 'symbolic_io_field',
            'Status display': 'status_display',
            'Switch': 'switch',
            'Slider': 'slider',
            'Rectangle': 'rectangle',
            'Line': 'line',
            'Circle': 'circle',
            'Image': 'image',
            'Faceplate': 'faceplate',
        }
        for prefix, t in type_map.items():
            if elem_name.startswith(prefix):
                elem_type = t
                break

        # Try to extract coordinate data from the binary region after the element name
        # The pattern in TIA RDF: element name, then 4-byte fields for properties
        region = data[elem_start:min(elem_start + 200, len(data))]

        # Look for coordinate-like patterns (small integers that could be pixel positions)
        coords = []
        for i in range(len(elem_name) + 4, min(len(region), 80)):
            # Try reading as 4-byte little-endian values
            if i + 4 <= len(region):
                val = struct.unpack_from('<f', region, i)[0]
                if 0 < val < 2000 and val == int(val):  # integer-like pixel value
                    coords.append(int(val))

        # Extract tag references from the binary region near this element
        tag_refs = []
        region_str = region.decode('utf-8', errors='ignore')
        for tm in re.finditer(r'(DB_[A-Z_][A-Z_0-9 .\-]{5,}|HOPPER_[A-Z_0-9]+|DataToerenw_[A-Z_0-9]+)', region_str):
            tag_refs.append(tm.group(1).strip())

        elements.append({
            'name': elem_name,
            'type': elem_type,
            'events': [],
            'tag_references_in_region': tag_refs,
        })

    return elements


def parse_screen_rdf(filepath, hmi_tags, ocr_data):
    """Parse a single screen RDF file and extract all HMI data."""
    with open(filepath, 'rb') as f:
        data = f.read()

    if len(data) < 10:
        return None

    # --- Screen name ---
    all_strings = []
    for m in re.finditer(rb'[\x20-\x7e]{3,}', data):
        s = m.group().decode('ascii', errors='ignore').strip()
        if s:
            all_strings.append(s)

    screen_name = None
    known_screens = [
        'START_SCHERM', 'WERKSCHERM', 'MANUEEL', 'AUTO_STAND',
        'INSTELLINGEN', 'ALARM', 'ALARMS_1', 'ALARMS_2',
        'MANUEEL_BANDEN_MIXAS', 'MANUEEL_VLOER',
        '00_ALGEMEEN', '01_ALGEMEEN', '01_HOPPER_OPVOER_BANDEN',
        '02_HOPPER_SPITASSEN', '03_HOPPER_VLOER',
        'HOPPER02_VETPOMP', 'VERLICHTING_HOPPER', 'BELTS_STATUS',
        '01_HOPPER_LINIAAL_INSTELLING', '02_HOPPER_LINIAAL01',
        'TIJD', 'WEIDMULLER', 'WEIDMULLER_WEGING',
        '05_HOPPER_SPITAS_DRUK', '08_HOPPER_LINIAAL_STATUS',
        'SYS_ANALOOG_STROOM_MIXAS', 'BEWAKINGEN', 'TOERENWACHTER',
        'TALEN', 'PROCES_1', 'PROCES_2', 'WATER',
        '01 ALARMS_GENERAL', '02 ALARMS HOPPER',
    ]
    for s in all_strings:
        for known in known_screens:
            if known in s and len(s) < len(known) + 5:
                screen_name = known
                break
        if screen_name:
            break
    if not screen_name:
        for s in all_strings:
            if re.match(r'^[A-Z][A-Z0-9_]{3,}$', s) and 'LAYER' not in s.upper() and 'Module' not in s:
                screen_name = s
                break

    # --- Elements ---
    elements = extract_element_details(data)

    # --- JavaScript functions ---
    js_functions = extract_js_functions(data)
    all_js_code = '\n'.join(f['body'] for f in js_functions)
    plc_tags_from_js = extract_plc_tags(all_js_code)
    navigations = extract_screen_navigation(all_js_code)

    # --- Link JS functions to elements ---
    for func in js_functions:
        func_name = func['name']
        func_tags = extract_plc_tags(func['body'])
        func_navs = extract_screen_navigation(func['body'])

        for elem in elements:
            elem_key = elem['name'].replace(' ', '_')
            if func_name.startswith(elem_key + '_'):
                # Determine event type from function suffix
                event_type = 'unknown'
                if '_OnUp' in func_name:
                    event_type = 'OnUp (release)'
                elif '_OnDown' in func_name:
                    event_type = 'OnDown (press)'
                elif '_OnLoaded' in func_name:
                    event_type = 'OnLoaded (screen init)'
                elif '_OnClick' in func_name:
                    event_type = 'OnClick'
                elif '_OnChange' in func_name:
                    event_type = 'OnChange'
                elif '_OnMouseEnter' in func_name:
                    event_type = 'OnMouseEnter'
                elif '_OnMouseLeave' in func_name:
                    event_type = 'OnMouseLeave'

                # Resolve PLC tag details from HMITags.xlsx
                resolved_tags = []
                for tag_name in func_tags:
                    tag_detail = hmi_tags.get(tag_name)
                    if tag_detail:
                        resolved_tags.append(tag_detail)
                    else:
                        resolved_tags.append({
                            'hmi_tag': tag_name,
                            'plc_tag': '(not found in HMITags.xlsx)',
                        })

                elem['events'].append({
                    'function': func_name,
                    'event_type': event_type,
                    'code': func['body'],
                    'plc_tags': func_tags,
                    'resolved_plc_tags': resolved_tags,
                    'navigates_to': func_navs,
                })

    # --- OnLoaded screen event ---
    on_loaded = None
    for func in js_functions:
        if 'OnLoaded' in func['name']:
            on_loaded = func['body']

    # --- Enrich elements with HMITag data ---
    for elem in elements:
        # Collect all PLC tags referenced by this element
        all_elem_tags = set()
        for ev in elem.get('events', []):
            for t in ev.get('plc_tags', []):
                all_elem_tags.add(t)
        for t in elem.get('tag_references_in_region', []):
            all_elem_tags.add(t)

        # Resolve to full tag details
        elem['hmi_tags'] = []
        for tag_name in sorted(all_elem_tags):
            tag_detail = hmi_tags.get(tag_name)
            if tag_detail:
                elem['hmi_tags'].append(tag_detail)
            else:
                elem['hmi_tags'].append({
                    'hmi_tag': tag_name,
                    'plc_tag': '(not found in HMITags.xlsx)',
                    'note': 'Referenced in JS but not in tag table',
                })

        # Classify element as input/output/display
        elem['io_role'] = classify_io_role(elem)

    # --- Enrich with OCR data ---
    ocr_screen_name = screen_name or ''
    ocr_elements = ocr_data.get(ocr_screen_name, [])
    if ocr_elements:
        # Try to match OCR elements to RDF elements by proximity
        for elem in elements:
            elem['ocr_match'] = find_ocr_match(elem, ocr_elements)

    # --- Layers ---
    layers = [s for s in all_strings if re.match(r'^Layer_\d+$', s)]

    return {
        'screen_name': screen_name or 'UNKNOWN',
        'file': os.path.basename(filepath),
        'elements': elements,
        'element_count': len(elements),
        'element_summary': summarize_elements(elements),
        'javascript_functions': [{'name': f['name'], 'body': f['body']} for f in js_functions],
        'plc_tags_referenced': plc_tags_from_js,
        'screen_navigations': navigations,
        'on_loaded_event': on_loaded,
        'layers': layers,
        'ocr_elements': ocr_elements if ocr_elements else None,
        'file_size': len(data),
    }


def classify_io_role(elem):
    """Classify element as input, output, display, navigation, or control."""
    events = elem.get('events', [])
    elem_type = elem.get('type', '')

    if elem_type == 'button':
        # Check if it navigates or controls
        for ev in events:
            if ev.get('navigates_to'):
                return 'navigation_button'
            if ev.get('plc_tags'):
                return 'control_button (writes to PLC)'
        return 'button'

    if elem_type == 'io_field':
        # IO fields can be input or output
        has_write = False
        has_read = False
        for ev in events:
            for tag in ev.get('plc_tags', []):
                if 'Set' in ev.get('code', '') or 'Write' in ev.get('code', ''):
                    has_write = True
                else:
                    has_read = True
        if has_write:
            return 'input (writes to PLC)'
        return 'output (reads from PLC)'

    if elem_type == 'symbolic_io_field':
        return 'selection_input'

    if elem_type == 'screen_window':
        return 'screen_container'

    if elem_type in ('text_display', 'rectangle', 'line', 'circle', 'image'):
        return 'display/static'

    if elem_type == 'bar_graph':
        return 'output (visualizes PLC value)'

    if elem_type == 'status_display':
        return 'output (PLC status)'

    if elem_type == 'switch':
        return 'input (writes to PLC)'

    return 'display/static'


def find_ocr_match(rdf_elem, ocr_elements):
    """Try to find matching OCR element(s) for an RDF element."""
    # Match by element type and approximate position
    type_map = {
        'button': ['button_green', 'button_red', 'button_blue', 'button_grey', 'button'],
        'io_field': ['input_field', 'display_field'],
        'text_display': ['label', 'title', 'display_field'],
        'screen_window': ['navigation'],
        'rectangle': ['label'],
        'circle': ['indicator_green', 'indicator_red', 'indicator_orange'],
    }

    elem_type = rdf_elem.get('type', '')
    candidate_types = type_map.get(elem_type, [])

    matches = []
    for ocr_elem in ocr_elements:
        if ocr_elem.get('type_ocr', '') in candidate_types:
            matches.append({
                'ocr_text': ocr_elem.get('text', ''),
                'ocr_type': ocr_elem.get('type_ocr', ''),
                'position': ocr_elem.get('center', {}),
                'bbox': ocr_elem.get('bbox', {}),
                'color': ocr_elem.get('color_region'),
                'confidence': ocr_elem.get('confidence', 0),
            })

    return matches if matches else None


def summarize_elements(elements):
    """Create a summary of element types on a screen."""
    summary = defaultdict(int)
    for e in elements:
        summary[e.get('type', 'unknown')] += 1
        if e.get('events'):
            summary['with_events'] += 1
        if e.get('hmi_tags'):
            summary['with_tag_bindings'] += 1
        io = e.get('io_role', '')
        if 'input' in io:
            summary['inputs'] += 1
        elif 'output' in io or 'display' in io:
            summary['outputs'] += 1
        elif 'control' in io or 'navigation' in io:
            summary['controls'] += 1
    return dict(summary)


# ---------------------------------------------------------------------------
# HMI instance discovery
# ---------------------------------------------------------------------------
def list_hmi_instances(project_path):
    """List all HMI instances found in the project with device info."""
    instances = {}
    base_dir = os.path.join(project_path, 'IM', 'HMI', 'I')
    if not os.path.exists(base_dir):
        return instances

    for entry in os.listdir(base_dir):
        entry_path = os.path.join(base_dir, entry)
        if not os.path.isdir(entry_path):
            continue
        for sub in ['Saved', 'Context']:
            dt_path = os.path.join(entry_path, sub, 'DownloadTask.xml')
            if os.path.exists(dt_path):
                try:
                    with open(dt_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    info = {}
                    for field in ['ESDeviceName', 'RtProjectFolderName', 'DeviceType',
                                  'DeviceVersion', 'CreationTime']:
                        m = re.search(rf'{field}="([^"]*)"', content)
                        if m:
                            info[field] = m.group(1)
                    screen_count = 0
                    screens_dir = os.path.join(entry_path, sub, 'screens')
                    if os.path.exists(screens_dir):
                        screen_count = len([f for f in os.listdir(screens_dir)
                                            if f.startswith('screen_') and f.endswith('.rdf')])
                    instances[entry] = {
                        'source': sub,
                        'device_type': info.get('DeviceType', 'Unknown'),
                        'device_version': info.get('DeviceVersion', '?'),
                        'device_name': info.get('RtProjectFolderName', 'Unknown'),
                        'es_device': info.get('ESDeviceName', '?'),
                        'creation_time': info.get('CreationTime', '?'),
                        'screen_count': screen_count,
                    }
                except Exception:
                    instances[entry] = {'source': sub, 'error': 'could not parse'}
                break
    return instances


def find_screen_files(project_path, instance_id=None):
    """Find screen RDF files, optionally filtered by instance."""
    base_dir = os.path.join(project_path, 'IM', 'HMI', 'I')
    if not os.path.exists(base_dir):
        return []

    if instance_id is not None:
        instance_dirs = [os.path.join(base_dir, str(instance_id))]
    else:
        instances = list_hmi_instances(project_path)
        if instances:
            latest = max(
                ((k, v) for k, v in instances.items() if 'creation_time' in v),
                key=lambda kv: kv[1].get('creation_time', ''),
                default=(None, {})
            )
            if latest[0]:
                instance_dirs = [os.path.join(base_dir, latest[0])]
                print(f"Auto-selected instance {latest[0]} (newest: {latest[1].get('creation_time','?')})")
            else:
                instance_dirs = [base_dir]
        else:
            instance_dirs = [base_dir]

    screen_files = set()
    for search_dir in instance_dirs:
        if not os.path.exists(search_dir):
            continue
        for root, dirs, files in os.walk(search_dir):
            for f in files:
                if f.startswith('screen_') and f.endswith('.rdf'):
                    full_path = os.path.join(root, f)
                    if 'Saved' in full_path:
                        screen_files.add(full_path)

    return sorted(screen_files)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    default_project_dir = os.path.dirname(os.path.abspath(__file__))

    parser = argparse.ArgumentParser(
        description='TIA Portal HMI Full Data Extractor - RDF + Tags + OCR',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('project_path', help='Path to TIA Portal project directory')
    parser.add_argument('--output', '-o', default=None,
                        help='Output JSON path (default: Doc_OUTPUT/hmi_offline_data.json)')
    parser.add_argument('--instance', '-i', type=int, default=None,
                        help='HMI instance ID (auto-detects latest)')
    parser.add_argument('--tags', default=None,
                        help='HMITags.xlsx path (default: DATA_HMI/HMITags.xlsx)')
    parser.add_argument('--ocr', default=None,
                        help='OCR all_elements.json path (default: Doc_OUTPUT/all_elements.json)')
    parser.add_argument('--list-hmis', action='store_true',
                        help='List all HMI instances and exit')
    parser.add_argument('--verbose', '-v', action='store_true')
    args = parser.parse_args()

    project_path = args.project_path.replace('\\', '/')

    if not os.path.exists(project_path):
        print(f"ERROR: Project path does not exist: {project_path}")
        sys.exit(1)

    # Resolve default paths relative to script location
    script_dir = default_project_dir
    output_path = args.output or os.path.join(script_dir, 'Doc_OUTPUT', 'hmi_offline_data.json')
    tags_path = args.tags or os.path.join(script_dir, 'DATA_HMI', 'HMITags.xlsx')
    ocr_path = args.ocr or os.path.join(script_dir, 'Doc_OUTPUT', 'all_elements.json')

    # --- List HMIs mode ---
    if args.list_hmis:
        instances = list_hmi_instances(project_path)
        if not instances:
            print("No HMI instances found.")
            sys.exit(1)
        print(f"Found {len(instances)} HMI instance(s):\n")
        for inst_id, info in sorted(instances.items()):
            print(f"  Instance I/{inst_id}:")
            for k, v in info.items():
                print(f"    {k}: {v}")
            print()
        return

    print(f"{'='*70}")
    print(f" TIA Portal HMI Full Data Extractor")
    print(f"{'='*70}")
    print(f" Project: {os.path.basename(project_path)}")

    # --- Load data sources ---
    print(f"\n--- Loading data sources ---")
    hmi_tags = load_hmi_tags(tags_path)
    ocr_data = load_ocr_data(ocr_path)

    # --- Find screen files ---
    print(f"\n--- Scanning for screens ---")
    screen_files = find_screen_files(project_path, instance_id=args.instance)
    print(f"Found {len(screen_files)} screen RDF files")

    if not screen_files:
        print("ERROR: No screen RDF files found.")
        sys.exit(1)

    # --- Parse screens ---
    screens = []
    errors = []

    for i, sf in enumerate(screen_files):
        try:
            result = parse_screen_rdf(sf, hmi_tags, ocr_data)
            if result:
                screens.append(result)
                if args.verbose:
                    name = result['screen_name']
                    n_elem = result['element_count']
                    n_tags = len(result['plc_tags_referenced'])
                    n_js = len(result['javascript_functions'])
                    n_nav = len(result['screen_navigations'])
                    summary = result.get('element_summary', {})
                    print(f"  [{i+1:2d}/{len(screen_files)}] {name:35s} | {n_elem:2d} elem | "
                          f"{n_js:2d} JS | {n_tags:2d} tags | {n_nav:2d} nav | {summary}")
        except Exception as e:
            errors.append({'file': sf, 'error': str(e)})
            if args.verbose:
                print(f"  [{i+1:2d}/{len(screen_files)}] ERROR: {sf} - {e}")

    # --- Build indexes ---
    nav_map = {}
    for s in screens:
        if s.get('screen_navigations'):
            nav_map[s['screen_name']] = s['screen_navigations']

    tag_index = defaultdict(list)
    for s in screens:
        for tag in s.get('plc_tags_referenced', []):
            tag_index[tag].append(s['screen_name'])

    # Resolve tag index entries with full PLC details
    resolved_tag_index = {}
    for tag_name, used_in in sorted(tag_index.items()):
        tag_detail = hmi_tags.get(tag_name)
        resolved_tag_index[tag_name] = {
            'plc_tag': tag_detail['plc_tag'] if tag_detail else '(not in tag table)',
            'data_type': tag_detail['data_type'] if tag_detail else '?',
            'connection': tag_detail['connection'] if tag_detail else '?',
            'used_in_screens': used_in,
        }

    # --- Compile output ---
    total_elements = sum(s['element_count'] for s in screens)
    total_events = sum(len([e for e in s['elements'] if e.get('events')]) for s in screens)
    total_tags = sum(len(e.get('hmi_tags', [])) for s in screens for e in s['elements'])

    # Get device info
    instance_info = None
    if screen_files:
        m = re.search(r'/I/(\d+)/', screen_files[0].replace('\\', '/'))
        if m:
            instances = list_hmi_instances(project_path)
            instance_info = instances.get(m.group(1), {})
            instance_info['instance_id'] = m.group(1)

    output = {
        'extraction_info': {
            'tool': 'extract_hmi_full.py',
            'source_project': os.path.basename(project_path),
            'timestamp': datetime.now().isoformat(),
            'data_sources': {
                'rdf_screens': len(screen_files),
                'hmi_tags_loaded': len(hmi_tags),
                'ocr_screens': len(ocr_data),
            },
            'screens_parsed': len(screens),
            'errors': len(errors),
        },
        'hmi_device': instance_info,
        'summary': {
            'total_screens': len(screens),
            'total_elements': total_elements,
            'total_elements_with_events': total_events,
            'total_tag_bindings': total_tags,
            'total_js_functions': sum(len(s['javascript_functions']) for s in screens),
            'total_unique_plc_tags': len(resolved_tag_index),
            'total_navigation_links': sum(len(v) for v in nav_map.values()),
        },
        'navigation_map': nav_map,
        'plc_tag_index': resolved_tag_index,
        'screens': screens,
    }

    if errors:
        output['errors'] = errors

    # --- Write output ---
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # --- Print report ---
    print(f"\n{'='*70}")
    print(f" EXTRACTION COMPLETE")
    print(f"{'='*70}")
    print(f"  Device:       {instance_info.get('device_name', '?') if instance_info else '?'}")
    print(f"  Screens:      {output['summary']['total_screens']}")
    print(f"  Elements:     {output['summary']['total_elements']}")
    print(f"  With events:  {output['summary']['total_elements_with_events']}")
    print(f"  Tag bindings: {output['summary']['total_tag_bindings']}")
    print(f"  JS functions: {output['summary']['total_js_functions']}")
    print(f"  Unique tags:  {output['summary']['total_unique_plc_tags']}")
    print(f"  Navigations:  {output['summary']['total_navigation_links']}")
    print(f"  Errors:       {len(errors)}")
    print(f"\n  Output: {output_path}")
    print(f"{'='*70}")

    # Print per-screen detail
    print(f"\n{'Screen':<36s} {'Elem':>5s} {'IO':>4s} {'Tags':>5s} {'JS':>4s} {'Nav':>4s} {'Summary'}")
    print('-' * 100)
    for s in screens:
        n_tags = sum(len(e.get('hmi_tags', [])) for e in s['elements'])
        n_inputs = sum(1 for e in s['elements'] if 'input' in e.get('io_role', ''))
        n_outputs = sum(1 for e in s['elements'] if 'output' in e.get('io_role', ''))
        n_ctrl = sum(1 for e in s['elements'] if 'control' in e.get('io_role', '') or 'navigation' in e.get('io_role', ''))
        io_str = f"{n_inputs}i/{n_outputs}o/{n_ctrl}c"
        summary_parts = [f"{k}:{v}" for k, v in sorted(s.get('element_summary', {}).items())
                         if k not in ('with_events', 'with_tag_bindings', 'inputs', 'outputs', 'controls')]
        print(f"  {s['screen_name']:<34s} {s['element_count']:>4d} {io_str:>8s} {n_tags:>4d} "
              f"{len(s['javascript_functions']):>4d} {len(s['screen_navigations']):>4d}  {', '.join(summary_parts)}")

    return output


if __name__ == '__main__':
    main()
